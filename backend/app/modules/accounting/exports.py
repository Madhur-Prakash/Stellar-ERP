"""Turning a balance sheet into a file someone can send to their accountant.

Two formats, one shared description of the document. :func:`_rows` flattens the statement
into the sequence of lines both writers walk, so the spreadsheet and the PDF cannot disagree
about what the balance sheet says - which is the failure that matters here. A number that
differs between two exports of the same report is worse than either export being ugly.

**Money is formatted from the `Decimal`, never from a float.** The API serialises amounts as
decimal strings precisely so no binary float is involved; re-parsing one into a `float` to
lay it out would reintroduce exactly what that choice avoids. The spreadsheet gets the
`Decimal` itself - openpyxl writes it as an exact numeric cell - and the PDF gets a string
built by grouping digits.
"""

from __future__ import annotations

import datetime as dt
import io
from dataclasses import dataclass
from decimal import Decimal
from typing import Final

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from app.modules.accounting.schemas import BalanceSheet

#: Excel's own number format, applied to every amount cell.
#:
#: A format string rather than a pre-formatted string, so the cell stays a *number*: the
#: recipient can sum a column, sort it, or change the currency symbol. Writing "₹1,234.00" as
#: text would look identical and be useless in a spreadsheet.
_MONEY_FORMAT: Final = "#,##0.00;[Red]-#,##0.00"

_HEADING: Final = Font(bold=True, size=12)
_LABEL: Final = Font(bold=True)
_TOTAL_RULE: Final = Border(top=Side(style="thin"), bottom=Side(style="double"))


@dataclass(frozen=True, slots=True)
class ExportRow:
    """One line of the rendered document.

    ``amount`` is None for a heading or a spacer, which is what lets both writers treat
    structure and figures uniformly instead of each inventing its own layout rules.
    """

    label: str
    amount: Decimal | None = None
    #: Indentation, carried through from `ReportLine.level`.
    level: int = 0
    #: A section heading - "Assets", "Liabilities".
    heading: bool = False
    #: A total, which both formats rule off.
    total: bool = False


def _rows(sheet: BalanceSheet, comparative: BalanceSheet | None) -> list[ExportRow]:
    """The document, as a flat list.

    Shared by both writers on purpose - see the module docstring. The comparative is not
    interleaved here: it is looked up by label at write time, because the two statements can
    legitimately hold different accounts (one opened mid-period) and zipping two lists by
    position would silently pair unrelated rows.
    """
    del comparative  # looked up by label below, not merged here

    rows: list[ExportRow] = []

    for title, lines, total, total_label in (
        ("Assets", sheet.assets, sheet.total_assets, "Total assets"),
        (
            "Liabilities",
            sheet.liabilities,
            sheet.total_liabilities,
            "Total liabilities",
        ),
        ("Equity", sheet.equity, sheet.total_equity, "Total equity"),
    ):
        rows.append(ExportRow(title, heading=True))
        for line in lines:
            rows.append(ExportRow(line.label, line.amount, level=line.level))

        if title == "Equity" and sheet.current_period_earnings:
            # Named explicitly rather than folded into retained earnings: until the year is
            # closed this profit is not there yet, and without the line the sheet does not
            # balance. Anyone checking the arithmetic needs to see it.
            rows.append(
                ExportRow(
                    "Current period earnings",
                    sheet.current_period_earnings,
                    level=1,
                )
            )

        rows.append(ExportRow(total_label, total, total=True))
        rows.append(ExportRow(""))

    rows.append(
        ExportRow(
            "Liabilities and equity",
            sheet.total_liabilities + sheet.total_equity,
            total=True,
        )
    )
    return rows


def _group_indian(digits: str) -> str:
    """Digit grouping the Indian way - the last three, then twos.

    Hand-rolled because `locale` is process-global and depends on what the host happens to
    have installed, and because this has to match what the two clients already print. A
    figure grouped one way in the app and another in its own export is the kind of
    discrepancy that makes someone distrust the whole report.
    """
    if len(digits) <= 3:
        return digits
    head, tail = digits[:-3], digits[-3:]
    parts: list[str] = []
    while len(head) > 2:
        parts.insert(0, head[-2:])
        head = head[:-2]
    if head:
        parts.insert(0, head)
    return ",".join(parts) + "," + tail


def format_money(amount: Decimal, currency: str) -> str:
    """A figure for the PDF, where there is no cell format to lean on."""
    negative = amount < 0
    whole, _, frac = f"{abs(amount):.2f}".partition(".")
    grouped = _group_indian(whole) if currency == "INR" else f"{int(whole):,}"
    symbol = {"INR": "₹", "USD": "$", "EUR": "€", "GBP": "£"}.get(currency, f"{currency} ")
    return f"{'-' if negative else ''}{symbol}{grouped}.{frac}"


def _amount_by_label(sheet: BalanceSheet) -> dict[str, Decimal]:
    """Every figure in a statement, keyed by its label.

    How the comparative column is filled: by name, not by position. An account that did not
    exist at the earlier date simply has no entry, and its cell is left blank rather than
    borrowing the figure from whichever row happened to sit at the same index.
    """
    found: dict[str, Decimal] = {}
    for line in (*sheet.assets, *sheet.liabilities, *sheet.equity):
        found[line.label] = line.amount
    found["Total assets"] = sheet.total_assets
    found["Total liabilities"] = sheet.total_liabilities
    found["Total equity"] = sheet.total_equity
    found["Current period earnings"] = sheet.current_period_earnings
    found["Liabilities and equity"] = sheet.total_liabilities + sheet.total_equity
    return found


def to_xlsx(
    sheet: BalanceSheet,
    *,
    organization: str,
    currency: str,
    comparative: BalanceSheet | None = None,
) -> bytes:
    """The balance sheet as a real `.xlsx`.

    Amounts are written as numbers with a currency *format*, so the recipient gets a
    spreadsheet they can work in rather than a picture of one.
    """
    book = Workbook()
    page = book.active
    assert page is not None  # noqa: S101 - a new Workbook always has one sheet
    page.title = "Balance sheet"

    page["A1"] = organization
    page["A1"].font = Font(bold=True, size=14)
    page["A2"] = "Balance sheet"
    page["A2"].font = _HEADING
    page["A3"] = f"As at {sheet.as_of.isoformat()}"
    page["A3"].font = Font(italic=True, color="6B7280")
    if not sheet.is_balanced:
        # Surfaced rather than hidden, for the same reason the API returns `is_balanced`: a
        # statement that does not balance is a fact about the books, and an export that
        # quietly omitted it would be the most misleading version of this file.
        page["A4"] = "Does not balance - assets do not equal liabilities plus equity."
        page["A4"].font = Font(bold=True, color="B91C1C")

    header = 6
    page.cell(row=header, column=1, value="").font = _LABEL
    page.cell(row=header, column=2, value=sheet.as_of.isoformat()).font = _LABEL
    page.cell(row=header, column=2).alignment = Alignment(horizontal="right")
    if comparative is not None:
        page.cell(row=header, column=3, value=comparative.as_of.isoformat()).font = _LABEL
        page.cell(row=header, column=3).alignment = Alignment(horizontal="right")

    prior = _amount_by_label(comparative) if comparative is not None else {}

    at = header + 1
    for row in _rows(sheet, comparative):
        if not row.label:
            at += 1
            continue

        label = page.cell(row=at, column=1, value=("    " * row.level) + row.label)
        if row.heading or row.total:
            label.font = _LABEL

        if row.amount is not None:
            cell = page.cell(row=at, column=2, value=row.amount)
            cell.number_format = _MONEY_FORMAT
            if row.total:
                cell.font = _LABEL
                cell.border = _TOTAL_RULE

            if comparative is not None and row.label in prior:
                was = page.cell(row=at, column=3, value=prior[row.label])
                was.number_format = _MONEY_FORMAT
                if row.total:
                    was.font = _LABEL
                    was.border = _TOTAL_RULE
        at += 1

    page.column_dimensions["A"].width = 46
    for column in range(2, 4 if comparative is not None else 3):
        page.column_dimensions[get_column_letter(column)].width = 18

    # Freeze below the column headers, so the labels stay put on a long chart of accounts.
    # A cell *reference* rather than the cell object: openpyxl accepts either at runtime, but
    # only the string is in the declared type, and a cast to silence that would be hiding the
    # mismatch rather than avoiding it.
    page.freeze_panes = f"A{header + 1}"

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def to_pdf(
    sheet: BalanceSheet,
    *,
    organization: str,
    currency: str,
    comparative: BalanceSheet | None = None,
) -> bytes:
    """The balance sheet as a PDF, for sending to someone who will only read it.

    **Helvetica, not a bundled font**, which means the rupee sign cannot be drawn: the core
    PDF fonts are Latin-1 only. `format_money` still yields "₹", so the symbol is replaced
    with the ISO code in this format and the currency is named in the header instead.
    Shipping a Unicode TTF would fix the glyph and add a megabyte to the image for one
    character.
    """
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    columns = 2 if comparative is not None else 1
    label_width = 190 - (columns * 35)

    def money(amount: Decimal) -> str:
        text = format_money(amount, currency)
        # Latin-1 only - see the docstring. Done here rather than in `format_money` so the
        # spreadsheet and the screen keep the real symbol.
        return text.encode("latin-1", "ignore").decode("latin-1") or f"{amount:.2f}"

    pdf.set_font("Helvetica", "B", 15)
    pdf.cell(0, 9, organization, new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 6, "Balance sheet", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(
        0,
        5,
        f"As at {sheet.as_of.isoformat()}   |   Amounts in {currency}",
        new_x="LMARGIN",
        new_y="NEXT",
    )
    if not sheet.is_balanced:
        pdf.set_text_color(185, 28, 28)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(
            0,
            5,
            "Does not balance - assets do not equal liabilities plus equity.",
            new_x="LMARGIN",
            new_y="NEXT",
        )
        pdf.set_text_color(0, 0, 0)
    pdf.ln(3)

    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(label_width, 6, "")
    pdf.cell(35, 6, sheet.as_of.isoformat(), align="R")
    if comparative is not None:
        pdf.cell(35, 6, comparative.as_of.isoformat(), align="R")
    pdf.ln(6)
    pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
    pdf.ln(1)

    prior = _amount_by_label(comparative) if comparative is not None else {}

    for row in _rows(sheet, comparative):
        if not row.label:
            pdf.ln(2)
            continue

        pdf.set_font("Helvetica", "B" if row.heading or row.total else "", 9)
        pdf.cell(label_width, 5.5, ("    " * row.level) + row.label)

        if row.amount is not None:
            pdf.cell(35, 5.5, money(row.amount), align="R")
            if comparative is not None:
                pdf.cell(
                    35,
                    5.5,
                    money(prior[row.label]) if row.label in prior else "",
                    align="R",
                )
        pdf.ln(5.5)

        if row.total:
            pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
            pdf.ln(1)

    out = pdf.output()
    return bytes(out)


def export_filename(kind: str, as_of: dt.date, extension: str) -> str:
    """A filename that sorts and says what it is: `balance-sheet-2026-03-31.xlsx`."""
    return f"{kind}-{as_of.isoformat()}.{extension}"
